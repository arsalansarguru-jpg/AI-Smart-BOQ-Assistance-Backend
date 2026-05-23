import re
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.schemas import BoqLineItem

HEADERS = [
    "Item No",
    "Category",
    "Description",
    "Unit",
    "Quantity",
    "Rate",
    "Amount",
    "Remarks",
]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16)
META_FONT = Font(size=10, color="4B5563")
NUM_ALIGN = Alignment(horizontal="right", vertical="top")
TEXT_ALIGN = Alignment(vertical="top", wrap_text=True)

COLUMN_WIDTHS = [12, 20, 52, 10, 12, 14, 14, 28]


def safe_export_filename(source_filename: str) -> str:
    base = re.sub(r"\.[^.]+$", "", source_filename.strip() or "boq")
    base = re.sub(r'[<>:"/\\|?*]', "_", base)[:80]
    return f"{base}_BOQ_export.xlsx"


def build_boq_workbook(
    items: list[BoqLineItem],
    *,
    source_filename: str,
    project_name: str | None = None,
    summary: str | None = None,
) -> bytes:
    wb = Workbook()
    
    # Standard styles & fonts configuration
    FONT_FAMILY = "Segoe UI"
    EMERALD_ACCENT = "059669"
    
    TITLE_FONT = Font(name=FONT_FAMILY, size=13, bold=True, color="FFFFFF")
    HEADER_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
    BOLD_FONT = Font(name=FONT_FAMILY, size=10, bold=True)
    REGULAR_FONT = Font(name=FONT_FAMILY, size=10)
    
    META_HEADER_FONT = Font(name=FONT_FAMILY, size=11, bold=True, color=EMERALD_ACCENT)
    GRAND_TOTAL_FONT = Font(name=FONT_FAMILY, size=11, bold=True, color="000000")
    
    TITLE_FILL = PatternFill("solid", fgColor="1E293B")  # Slate 800
    HEADER_FILL = PatternFill("solid", fgColor="334155") # Slate 700
    SUB_FILL = PatternFill("solid", fgColor="F1F5F9")    # Slate 100
    ZEBRA_FILL = PatternFill("solid", fgColor="F8FAFC")  # Slate 50
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    double_bottom_border = Border(
        top=Side(style='thin', color='94A3B8'),
        bottom=Side(style='double', color='1E293B')  # Accounting double underline
    )
    
    # -------------------------------------------------------------
    # SHEET 1: TENDER SUMMARY
    # -------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Tender Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title Banner Block
    ws_summary.merge_cells("A2:C3")
    title_cell = ws_summary["A2"]
    title_cell.value = "VERTEX CONTRACTING & MEP SERVICES - ESTIMATION SUMMARY"
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Fill background for the entire merged title block
    for r in range(2, 4):
        for c in range(1, 4):
            ws_summary.cell(row=r, column=c).fill = TITLE_FILL
            
    # Metadata Block
    ws_summary.cell(row=5, column=1, value="PROJECT ESTIMATION METADATA").font = META_HEADER_FONT
    
    meta_items = [
        ("Project/Tender Name:", project_name or "N/A"),
        ("Source Document:", source_filename),
        ("Exported Date:", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Total Line Items:", len(items)),
        ("Estimation Status:", "AI Analyzed & Structured"),
    ]
    
    start_row = 6
    for idx, (label, val) in enumerate(meta_items):
        r = start_row + idx
        
        lbl_cell = ws_summary.cell(row=r, column=1, value=label)
        lbl_cell.font = BOLD_FONT
        lbl_cell.fill = SUB_FILL
        lbl_cell.border = thin_border
        
        val_cell = ws_summary.cell(row=r, column=2, value=val)
        val_cell.font = REGULAR_FONT
        val_cell.border = thin_border
        
        # Merge value cell across columns B and C
        ws_summary.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws_summary.cell(row=r, column=3).border = thin_border
        
    # Extract unique categories
    categories = []
    seen = set()
    for item in items:
        cat = item.category.strip() if item.category and item.category.strip() else "Miscellaneous"
        if cat not in seen:
            seen.add(cat)
            categories.append(cat)
            
    # Trade Cost Summary Block
    summary_start_row = 13
    ws_summary.cell(row=summary_start_row, column=1, value="BID COST SUMMARY BY TRADE CATEGORY").font = META_HEADER_FONT
    
    header_row = summary_start_row + 1
    summary_headers = ["Trade Category", "Total Line Items", "Estimated Cost"]
    for col_idx, h_text in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=header_row, column=col_idx, value=h_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
        cell.border = thin_border
        
    data_start = header_row + 1
    for idx, cat in enumerate(categories):
        r = data_start + idx
        
        # Category Name
        cat_cell = ws_summary.cell(row=r, column=1, value=cat)
        cat_cell.font = BOLD_FONT
        cat_cell.border = thin_border
        
        # COUNTIF formula referencing BOQ Details sheet (Column B is Category)
        count_cell = ws_summary.cell(row=r, column=2, value=f"=COUNTIF('BOQ Details'!B:B, A{r})")
        count_cell.font = REGULAR_FONT
        count_cell.alignment = Alignment(horizontal="right", vertical="center")
        count_cell.border = thin_border
        count_cell.number_format = "#,##0"
        
        # SUMIF formula referencing BOQ Details sheet (Column B is Category, Column G is Amount)
        cost_cell = ws_summary.cell(row=r, column=3, value=f"=SUMIF('BOQ Details'!B:B, A{r}, 'BOQ Details'!G:G)")
        cost_cell.font = BOLD_FONT
        cost_cell.alignment = Alignment(horizontal="right", vertical="center")
        cost_cell.border = thin_border
        cost_cell.number_format = '"₹"#,##0.00'
        
        # Soft Zebra layout for lists
        if idx % 2 == 1:
            cat_cell.fill = ZEBRA_FILL
            count_cell.fill = ZEBRA_FILL
            cost_cell.fill = ZEBRA_FILL
            
    # Grand Total row
    total_row = data_start + len(categories)
    
    total_lbl = ws_summary.cell(row=total_row, column=1, value="GRAND ESTIMATED TOTAL")
    total_lbl.font = GRAND_TOTAL_FONT
    total_lbl.fill = SUB_FILL
    total_lbl.border = double_bottom_border
    
    total_count = ws_summary.cell(row=total_row, column=2, value=f"=SUM(B{data_start}:B{total_row-1})")
    total_count.font = GRAND_TOTAL_FONT
    total_count.alignment = Alignment(horizontal="right", vertical="center")
    total_count.fill = SUB_FILL
    total_count.border = double_bottom_border
    total_count.number_format = "#,##0"
    
    total_cost = ws_summary.cell(row=total_row, column=3, value=f"=SUM(C{data_start}:C{total_row-1})")
    total_cost.font = GRAND_TOTAL_FONT
    total_cost.alignment = Alignment(horizontal="right", vertical="center")
    total_cost.fill = SUB_FILL
    total_cost.border = double_bottom_border
    total_cost.number_format = '"₹"#,##0.00'
    
    # Custom column spacing
    ws_summary.column_dimensions["A"].width = 34
    ws_summary.column_dimensions["B"].width = 18
    ws_summary.column_dimensions["C"].width = 22
    
    # -------------------------------------------------------------
    # SHEET 2: BOQ DETAILS
    # -------------------------------------------------------------
    ws_boq = wb.create_sheet(title="BOQ Details")
    ws_boq.views.sheetView[0].showGridLines = True
    
    # Title Banner Block
    ws_boq.merge_cells("A2:H3")
    boq_title_cell = ws_boq["A2"]
    boq_title_cell.value = "VERTEX CONTRACTING & MEP SERVICES - DETAILED ESTIMATION SHEET"
    boq_title_cell.font = TITLE_FONT
    boq_title_cell.fill = TITLE_FILL
    boq_title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for r in range(2, 4):
        for c in range(1, 9):
            ws_boq.cell(row=r, column=c).fill = TITLE_FILL
            
    # Subheading
    ws_boq.cell(row=5, column=1, value="DETAILED BILL OF QUANTITIES (BOQ)").font = META_HEADER_FONT
    
    # Headers
    header_row_boq = 7
    headers_boq = ["Item No", "Category", "Description", "Unit", "Quantity", "Rate", "Amount", "Remarks"]
    for col_idx, h_text in enumerate(headers_boq, start=1):
        cell = ws_boq.cell(row=header_row_boq, column=col_idx, value=h_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center" if col_idx in (1, 4, 5, 6, 7) else "left", vertical="center")
        cell.border = thin_border
        
    # Data list
    data_start_boq = 8
    for idx, item in enumerate(items):
        r = data_start_boq + idx
        
        # Standardize categories and numeric defaults
        cat = item.category.strip() if item.category and item.category.strip() else "Miscellaneous"
        qty = item.quantity if item.quantity is not None else 0.0
        rate = item.rate if item.rate is not None else 0.0
        
        # Write values
        c_item = ws_boq.cell(row=r, column=1, value=item.item_no)
        c_item.alignment = Alignment(horizontal="center", vertical="top")
        
        c_cat = ws_boq.cell(row=r, column=2, value=cat)
        c_cat.alignment = Alignment(vertical="top")
        c_cat.font = Font(name=FONT_FAMILY, size=10, bold=True, color="475569")
        
        c_desc = ws_boq.cell(row=r, column=3, value=item.description)
        c_desc.alignment = Alignment(vertical="top", wrap_text=True)
        
        c_unit = ws_boq.cell(row=r, column=4, value=item.unit)
        c_unit.alignment = Alignment(horizontal="center", vertical="top")
        
        c_qty = ws_boq.cell(row=r, column=5, value=qty)
        c_qty.alignment = Alignment(horizontal="right", vertical="top")
        c_qty.number_format = "#,##0.00"
        
        c_rate = ws_boq.cell(row=r, column=6, value=rate)
        c_rate.alignment = Alignment(horizontal="right", vertical="top")
        c_rate.number_format = '"₹"#,##0.00'
        
        # Dynamic quantity * rate multiplication formula!
        c_amount = ws_boq.cell(row=r, column=7, value=f"=E{r}*F{r}")
        c_amount.alignment = Alignment(horizontal="right", vertical="top")
        c_amount.font = BOLD_FONT
        c_amount.number_format = '"₹"#,##0.00'
        
        c_remarks = ws_boq.cell(row=r, column=8, value=item.remarks)
        c_remarks.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Apply zebra coloring and thin borders to columns
        for col_idx in range(1, 9):
            cell = ws_boq.cell(row=r, column=col_idx)
            if col_idx not in (2, 7): # preserve customized fonts for category / amount
                cell.font = REGULAR_FONT
            cell.border = thin_border
            if idx % 2 == 1:
                cell.fill = ZEBRA_FILL
                
    # Grand Total row at the bottom of Details
    total_row_boq = data_start_boq + len(items)
    
    lbl_total = ws_boq.cell(row=total_row_boq, column=6, value="TOTAL")
    lbl_total.font = GRAND_TOTAL_FONT
    lbl_total.alignment = Alignment(horizontal="right", vertical="center")
    lbl_total.fill = SUB_FILL
    lbl_total.border = double_bottom_border
    
    val_total = ws_boq.cell(row=total_row_boq, column=7, value=f"=SUM(G{data_start_boq}:G{total_row_boq-1})")
    val_total.font = GRAND_TOTAL_FONT
    val_total.alignment = Alignment(horizontal="right", vertical="center")
    val_total.fill = SUB_FILL
    val_total.border = double_bottom_border
    val_total.number_format = '"₹"#,##0.00'
    
    # Clean double border lines for the rest of total columns
    for col_idx in range(1, 9):
        if col_idx not in (6, 7):
            cell = ws_boq.cell(row=total_row_boq, column=col_idx)
            cell.border = double_bottom_border
            cell.fill = SUB_FILL
            
    # Set proper dimensions for scannability
    ws_boq.column_dimensions["A"].width = 12
    ws_boq.column_dimensions["B"].width = 24
    ws_boq.column_dimensions["C"].width = 56
    ws_boq.column_dimensions["D"].width = 10
    ws_boq.column_dimensions["E"].width = 14
    ws_boq.column_dimensions["F"].width = 14
    ws_boq.column_dimensions["G"].width = 16
    ws_boq.column_dimensions["H"].width = 28
    
    # Freeze header rows in BOQ Details
    ws_boq.freeze_panes = ws_boq.cell(row=data_start_boq, column=1)
    
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
