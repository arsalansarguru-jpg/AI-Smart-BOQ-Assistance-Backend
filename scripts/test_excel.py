import os
import sys

# Ensure backend directory is in the path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.export.excel import build_boq_workbook
from app.models.schemas import BoqLineItem

def test_generation():
    items = [
        BoqLineItem(
            item_no="1.1",
            category="Copper Piping",
            description="Supply and installation of 22mm copper piping for VRF liquid line, including standard insulation and supports.",
            unit="m",
            quantity=45.0,
            rate=18.50,
            amount=832.50,
            remarks="Liquid line"
        ),
        BoqLineItem(
            item_no="1.2",
            category="Copper Piping",
            description="Supply and installation of 28mm copper piping for VRF suction line, including thick elastomeric foam insulation.",
            unit="m",
            quantity=45.0,
            rate=24.00,
            amount=1080.00,
            remarks="Suction line"
        ),
        BoqLineItem(
            item_no="2.1",
            category="HVAC Ducting",
            description="Fabrication and installation of GI ducting (24g) including neoprene gaskets, hangers, and dampers.",
            unit="m2",
            quantity=120.0,
            rate=45.00,
            amount=5400.00,
            remarks="Exhaust duct"
        ),
        BoqLineItem(
            item_no="2.2",
            category="HVAC Ducting",
            description="Fabrication and installation of GI ducting (22g) for main supply air ducts, including thermal insulation.",
            unit="m2",
            quantity=85.0,
            rate=55.00,
            amount=4675.00,
            remarks="Supply air duct"
        ),
        BoqLineItem(
            item_no="3.1",
            category="Electrical Wiring",
            description="Laying of 3-core 2.5 sqmm copper XLPE cable in existing conduit, including glands and terminations.",
            unit="m",
            quantity=300.0,
            rate=4.50,
            amount=1350.00,
            remarks="VRF power feed"
        ),
        BoqLineItem(
            item_no="3.2",
            category="Electrical Wiring",
            description="Laying of 2-core 1.5 sqmm shielded communication cable for indoor-outdoor interconnectivity.",
            unit="m",
            quantity=250.0,
            rate=3.20,
            amount=800.00,
            remarks="Control cable"
        ),
    ]

    print("Generating workbook...")
    wb_bytes = build_boq_workbook(
        items=items,
        source_filename="VRF_AC_System_Estimation.pdf",
        project_name="Commercial Tower VRF AC Installation",
        summary="Complete VRF Air Conditioning installation including piping, ducting, and power connections."
    )

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_boq_output.xlsx")
    with open(out_path, "wb") as f:
        f.write(wb_bytes)
        
    print(f"Workbook successfully saved to: {out_path}")

    # Verify the contents using openpyxl
    from openpyxl import load_workbook
    wb = load_workbook(out_path, data_only=False) # Load with formulas, not evaluated values
    
    print("\n--- SHEET VERIFICATION ---")
    print(f"Sheets found: {wb.sheetnames}")
    
    # 1. Verify Tender Summary Sheet
    ws_summary = wb["Tender Summary"]
    print("\nChecking 'Tender Summary' sheet:")
    print(f"Title Block (A2): {ws_summary['A2'].value}")
    print(f"Project Metadata Name (B6): {ws_summary['B6'].value}")
    print(f"Total Line Items Count (B9): {ws_summary['B9'].value}")
    
    print("\nTrade Categories List:")
    # We should have categories in column A starting at row 15
    for r in range(15, 18):
        cat = ws_summary.cell(row=r, column=1).value
        count_formula = ws_summary.cell(row=r, column=2).value
        cost_formula = ws_summary.cell(row=r, column=3).value
        print(f"  Row {r}: Category={cat} | CountFormula={count_formula} | CostFormula={cost_formula}")
        
    # Grand Total row
    total_row = 18
    print(f"Grand Total Row {total_row}:")
    print(f"  Count sum formula: {ws_summary.cell(row=total_row, column=2).value}")
    print(f"  Cost sum formula: {ws_summary.cell(row=total_row, column=3).value}")
    
    # 2. Verify BOQ Details Sheet
    ws_boq = wb["BOQ Details"]
    print("\nChecking 'BOQ Details' sheet:")
    print(f"Title Block (A2): {ws_boq['A2'].value}")
    
    print("\nFirst two detail line items:")
    for r in range(8, 10):
        item_no = ws_boq.cell(row=r, column=1).value
        desc = ws_boq.cell(row=r, column=3).value[:30] + "..."
        qty = ws_boq.cell(row=r, column=5).value
        rate = ws_boq.cell(row=r, column=6).value
        amt_formula = ws_boq.cell(row=r, column=7).value
        print(f"  Row {r}: Item={item_no} | Desc={desc} | Qty={qty} | Rate={rate} | AmountFormula={amt_formula}")
        
    # BOQ total row
    boq_total_row = 14
    print(f"\nBOQ Grand Total Row {boq_total_row}:")
    print(f"  Total label: {ws_boq.cell(row=boq_total_row, column=6).value}")
    print(f"  Total sum formula: {ws_boq.cell(row=boq_total_row, column=7).value}")
    
    print("\nVerification completed successfully!")

if __name__ == "__main__":
    test_generation()
